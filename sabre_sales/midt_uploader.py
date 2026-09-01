import frappe
import openpyxl
from frappe.utils import now


@frappe.whitelist()
def import_midt(upload_name):
    # Suppresses the per-agency doc_events hook while the import runs; the
    # single resolve_agency_links() pass at the end covers everything.
    frappe.flags.in_midt_import = True
    """Parse the Excel attached to a MIDT Upload doc and load it into MIDT Record.
    Replaces overlapping months so monthly re-uploads are safe."""
    upload = frappe.get_doc('MIDT Upload', upload_name)

    if not upload.midt_file:
        frappe.throw('Please attach the MIDT Excel file first.')

    allowed = ['System Manager', 'Sabre Head Of Sales', 'Sabre Account Manager']
    user_roles = frappe.get_roles(frappe.session.user)
    if not any(r in user_roles for r in allowed):
        frappe.throw('You are not allowed to import MIDT data.')

    file_doc = frappe.get_doc('File', {'file_url': upload.midt_file})
    path = file_doc.get_full_path()

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if 'Sheet1' in wb.sheetnames:
            ws = wb['Sheet1']
        else:
            ws = max(wb.worksheets, key=lambda s: s.max_row)

        rows = ws.iter_rows(values_only=True)
        header = next(rows)

        # Map columns by header name rather than position. The Jun25-May26 file
        # had 16 columns; the Aug24-Jul26 one has 13 (the three Share columns were
        # dropped, moving Total Bookings from index 15 to 12). Reading by name
        # survives that.
        col = {}
        for i, h in enumerate(header):
            key = str(h or '').strip().lower()
            if key and key not in col:
                col[key] = i

        def need(label):
            if label not in col:
                frappe.throw(
                    'Column "' + label + '" was not found in the file. '
                    'Columns present: ' + ', '.join(sorted(col.keys())))
            return col[label]

        i_month = need('month year')
        i_sc = need('sc code')
        i_scname = need('sc code name')
        i_pcc = need('pcc')
        i_sf = need('sf top account')
        i_iata = need('iata')
        i_iataname = need('iata name')
        i_supplier = need('supplier')
        i_sabre = need('sabre bookings')
        i_amadeus = need('amadeus bookings')
        i_travelport = need('travelport bookings')
        i_total = need('total bookings')

        data = []
        months = set()
        for r in rows:
            if not r or not r[i_month] or str(r[i_month]) == 'Overall':
                continue
            month = r[i_month].date() if hasattr(r[i_month], 'date') else r[i_month]
            months.add(str(month))
            data.append((month, r[i_sc], r[i_scname], r[i_pcc], r[i_iata],
                         r[i_iataname], r[i_supplier],
                         int(r[i_sabre] or 0), int(r[i_amadeus] or 0),
                         int(r[i_travelport] or 0), int(r[i_total] or 0),
                         r[i_sf]))

        if not data:
            frappe.throw('No data rows found in the file. Is this the correct MIDT export?')

        for m in months:
            frappe.db.sql("DELETE FROM `tabMIDT Record` WHERE month = %s", m)

        start = int(frappe.db.sql("""
            SELECT IFNULL(MAX(CAST(SUBSTRING(name, 6) AS UNSIGNED)), 0)
            FROM `tabMIDT Record` WHERE name LIKE 'MIDT-%'
        """)[0][0])

        ts = now()
        user = frappe.session.user
        fields = ['name', 'owner', 'creation', 'modified', 'modified_by', 'docstatus', 'idx',
                  'month', 'sc_code', 'sc_code_name', 'pcc', 'iata', 'iata_name', 'supplier',
                  'sabre_bookings', 'amadeus_bookings', 'travelport_bookings', 'total_bookings',
                  'sf_top_account']
        values = []
        for i, d in enumerate(data):
            name = 'MIDT-' + str(start + i + 1).zfill(7)
            values.append((name, user, ts, ts, user, 0, 0) + d)

        chunk = 5000
        total_chunks = (len(values) + chunk - 1) // chunk
        for ci in range(total_chunks):
            frappe.db.bulk_insert('MIDT Record', fields,
                                  values[ci * chunk:(ci + 1) * chunk], chunk_size=chunk)
            frappe.publish_progress(
                (ci + 1) * 100.0 / total_chunks,
                title='Importing MIDT data',
                description=str(min((ci + 1) * chunk, len(values))) + ' of ' + str(len(values)) + ' records'
            )

        new_agencies = _sync_agencies()
        new_sf_agencies = _sync_placeholder_accounts()
        sc_updated = _refresh_sc_codes()
        resolution = resolve_agency_links()

        month_list = ', '.join(sorted(months))
        frappe.db.set_value('MIDT Upload', upload_name, {
            'status': 'Imported',
            'records_imported': len(values),
            'import_log': 'Imported ' + str(len(values)) + ' rows for months: ' +
                          month_list + ' | New agencies created: ' + str(new_agencies) +
                          ' | New SF-account agencies: ' + str(new_sf_agencies) +
                          ' | SC codes refreshed: ' + str(sc_updated) +
                          ' | Rows linked to an agency: ' + str(resolution['resolved']) +
                          ' | Unresolved: ' + str(resolution['unresolved']),
        }, update_modified=False)
        frappe.db.commit()

        warn = ''
        if resolution['unresolved']:
            warn = ('<br><br><b>Warning:</b> ' + str(resolution['unresolved']) +
                    ' rows (' + str(resolution['unresolved_bookings']) +
                    ' bookings) could not be linked to an agency and will not '
                    'appear in any agency figure.')

        _notify(upload, 'MIDT import finished',
                'Imported ' + str(len(values)) + ' rows for months: ' + month_list +
                '. New agencies created: ' + str(new_agencies) + '.' + warn,
                ok=True)

        return {'ok': True, 'records': len(values), 'months': sorted(months),
                'new_agencies': new_agencies}

    except Exception as e:
        frappe.db.rollback()
        frappe.db.set_value('MIDT Upload', upload_name, {
            'status': 'Failed',
            'import_log': str(e)[:500],
        }, update_modified=False)
        frappe.db.commit()
        _notify(upload, 'MIDT import failed', str(e)[:400], ok=False)
        raise


def _sync_agencies():
    """Create Sabre Agency records for any IATA present in MIDT but not yet an agency.
    Runs after every import so new agencies appear automatically each month."""
    rows = frappe.db.sql("""
        SELECT iata,
               (SELECT r3.iata_name FROM `tabMIDT Record` r3
                WHERE r3.iata = r.iata AND r3.iata_name IS NOT NULL
                  AND r3.iata_name NOT IN ('-', 'Overall')
                GROUP BY r3.iata_name ORDER BY SUM(r3.total_bookings) DESC LIMIT 1) AS iata_name,
               (SELECT r2.pcc FROM `tabMIDT Record` r2
                WHERE r2.iata = r.iata
                GROUP BY r2.pcc ORDER BY SUM(r2.total_bookings) DESC LIMIT 1) AS main_pcc
        FROM `tabMIDT Record` r
        WHERE iata IS NOT NULL AND iata != ''
          AND iata NOT IN ('1111111', '9999999')
          AND iata_name IS NOT NULL AND iata_name NOT IN ('-', 'Overall')
        GROUP BY iata
    """, as_dict=True)

    existing = set(x[0] for x in frappe.db.sql("SELECT iata FROM `tabSabre Agency`"))
    used_names = set(x[0] for x in frappe.db.sql("SELECT name FROM `tabSabre Agency`"))

    ts = now()
    user = frappe.session.user
    fields = ['name', 'owner', 'creation', 'modified', 'modified_by', 'docstatus', 'idx',
              'agency_name', 'iata', 'pcc', 'current_gds']
    values = []
    for a in rows:
        if a.iata in existing:
            continue
        name = (a.iata_name or a.iata).strip()[:130]
        if name in used_names:
            name = name + ' - ' + a.iata
        used_names.add(name)
        values.append((name, user, ts, ts, user, 0, 0,
                       name, a.iata, a.main_pcc or '', ''))

    if values:
        frappe.db.bulk_insert('Sabre Agency', fields, values, chunk_size=1000)
    return len(values)


def _notify(upload, subject, body, ok=True):
    """Tell whoever started the import how it went - it runs in the background,
    so nobody is watching the screen when it finishes.

    Two channels: a Notification Log entry that persists, and a realtime push
    so a popup appears straight away if they still have the page open.
    """
    user = upload.owner or frappe.session.user
    if not user or user in ('Administrator', 'Guest'):
        user = frappe.session.user

    try:
        frappe.get_doc({
            'doctype': 'Notification Log',
            'for_user': user,
            'type': 'Alert',
            'document_type': 'MIDT Upload',
            'document_name': upload.name,
            'subject': subject,
            'email_content': body,
        }).insert(ignore_permissions=True)
        frappe.db.commit()
    except Exception:
        # never let the notification break the import
        frappe.db.rollback()

    try:
        frappe.publish_realtime(
            'midt_import_update',
            {
                'upload': upload.name,
                'subject': subject,
                'body': body,
                'ok': 1 if ok else 0,
            },
            user=user,
        )
    except Exception:
        pass


@frappe.whitelist()
def enqueue_import_midt(upload_name):
    """Queue the import instead of running it inside the web request.

    The full file takes around 2-3 minutes, which is longer than the gateway
    will wait, so running it inline made the UI report a failure for an import
    that had actually succeeded.
    """
    upload = frappe.get_doc('MIDT Upload', upload_name)

    if not upload.midt_file:
        frappe.throw('Please attach the MIDT Excel file first.')

    allowed = ['System Manager', 'Sabre Head Of Sales', 'Sabre Account Manager']
    if not any(r in frappe.get_roles(frappe.session.user) for r in allowed):
        frappe.throw('You are not allowed to import MIDT data.')

    frappe.db.set_value('MIDT Upload', upload_name, {
        'status': 'Pending',
        'import_log': 'Queued at ' + now() + '. This usually takes 2-3 minutes; '
                      'you will get a notification when it finishes.',
    }, update_modified=False)
    frappe.db.commit()

    frappe.enqueue(
        'sabre_sales.midt_uploader.import_midt',
        queue='long',
        timeout=1800,
        upload_name=upload_name,
    )

    return {'queued': True}


def _refresh_sc_codes():
    """Keep Sabre Agency.sc_code / sc_code_name in step with MIDT.

    The dominant code is the one carrying the most bookings for that agency.
    Agencies with an SF Top Account are scoped to the placeholder IATAs, the
    rest to their own IATA - the same rule the panel and reports use.
    Returns how many rows were touched, or -1 if it could not run.
    """
    try:
        # 1. agencies keyed on their own IATA
        frappe.db.sql("""
            UPDATE `tabSabre Agency` a
            JOIN (
                SELECT iata, sc_code, sc_code_name,
                       ROW_NUMBER() OVER (
                           PARTITION BY iata ORDER BY SUM(total_bookings) DESC
                       ) AS rn
                FROM `tabMIDT Record`
                WHERE IFNULL(iata, '') != ''
                GROUP BY iata, sc_code, sc_code_name
            ) x ON x.iata = a.iata AND x.rn = 1
            SET a.sc_code = x.sc_code,
                a.sc_code_name = x.sc_code_name
            WHERE IFNULL(a.sf_top_account, '') = ''
              AND IFNULL(a.iata, '') != ''
        """)
        by_iata = frappe.db._cursor.rowcount if hasattr(frappe.db, '_cursor') else 0

        # 2. agencies keyed on SF Top Account (the placeholder-IATA companies)
        frappe.db.sql("""
            UPDATE `tabSabre Agency` a
            JOIN (
                SELECT sf_top_account, sc_code, sc_code_name,
                       ROW_NUMBER() OVER (
                           PARTITION BY sf_top_account ORDER BY SUM(total_bookings) DESC
                       ) AS rn
                FROM `tabMIDT Record`
                WHERE iata IN ('1111111', '9999999')
                  AND IFNULL(sf_top_account, '') != ''
                GROUP BY sf_top_account, sc_code, sc_code_name
            ) x ON x.sf_top_account = a.sf_top_account AND x.rn = 1
            SET a.sc_code = x.sc_code,
                a.sc_code_name = x.sc_code_name
            WHERE IFNULL(a.sf_top_account, '') != ''
        """)
        by_sf = frappe.db._cursor.rowcount if hasattr(frappe.db, '_cursor') else 0

        frappe.db.commit()
        return (by_iata or 0) + (by_sf or 0)
    except Exception:
        # never let this break the import - the data is already in
        frappe.db.rollback()
        return -1


def _sync_placeholder_accounts():
    """Give every company inside the placeholder IATAs its own agency card.

    `_sync_agencies` creates one card per IATA, which is useless for 1111111 and
    9999999 - those are placeholders shared by dozens of unrelated companies. The
    real identity there is SF Top Account, so each one gets its own card scoped by
    `sf_top_account` instead of by IATA.

    Returns how many cards were created.
    """
    placeholders = ('1111111', '9999999')
    try:
        accounts = frappe.db.sql("""
            SELECT sf_top_account, SUM(total_bookings) AS total
            FROM `tabMIDT Record`
            WHERE iata IN %(ph)s AND IFNULL(sf_top_account, '') != ''
            GROUP BY sf_top_account
        """, {'ph': placeholders}, as_dict=True)

        mapped = set(x[0] for x in frappe.db.sql(
            "SELECT sf_top_account FROM `tabSabre Agency` "
            "WHERE IFNULL(sf_top_account, '') != ''"))

        existing_names = set(x[0] for x in frappe.db.sql(
            "SELECT name FROM `tabSabre Agency`"))

        created = 0
        for a in accounts:
            sf = (a.sf_top_account or '').strip()
            if not sf or sf in mapped:
                continue

            # If a card already carries this exact name, adopt it rather than
            # making a near-duplicate.
            if sf in existing_names:
                frappe.db.set_value('Sabre Agency', sf, 'sf_top_account', sf,
                                    update_modified=False)
                mapped.add(sf)
                continue

            pcc = sf.rsplit(' - ', 1)[1] if ' - ' in sf else ''
            doc = frappe.get_doc({
                'doctype': 'Sabre Agency',
                'agency_name': sf,
                'iata': '',
                'pcc': pcc,
                'sf_top_account': sf,
            })
            doc.insert(ignore_permissions=True)
            mapped.add(sf)
            existing_names.add(sf)
            created += 1

        frappe.db.commit()
        return created
    except Exception:
        frappe.db.rollback()
        return -1


def resolve_agency_links():
    """Point every MIDT Record at exactly one Sabre Agency.

    Done once here so no query ever has to re-derive agency scope. Rules, in
    order - later steps deliberately override earlier ones:

      1. plain      - the card on that IATA
      2. shared     - where a card matches (iata, sc_code), prefer it. Some
                      IATAs are consolidators carrying several companies, each
                      with its own SC Code.
      3. placeholder- IATAs 1111111 / 9999999 are shared by unrelated companies,
                      so those rows key on sf_top_account instead.

    Returns a dict of counts. Safe to run repeatedly.
    """
    placeholders = ('1111111', '9999999')

    # Start clean so a re-run cannot leave a stale link behind.
    frappe.db.sql("UPDATE `tabMIDT Record` SET agency = NULL")

    # 1. plain IATA
    frappe.db.sql("""
        UPDATE `tabMIDT Record` m
        JOIN `tabSabre Agency` a ON a.iata = m.iata
        SET m.agency = a.name
        WHERE IFNULL(a.iata, '') != '' AND m.iata NOT IN %(ph)s
    """, {'ph': placeholders})

    # 2. (iata, sc_code) wins where such a card exists
    frappe.db.sql("""
        UPDATE `tabMIDT Record` m
        JOIN `tabSabre Agency` a
          ON a.iata = m.iata AND a.sc_code = m.sc_code
        SET m.agency = a.name
        WHERE IFNULL(a.sc_code, '') NOT IN ('', '-', 'UNK')
          AND m.iata NOT IN %(ph)s
    """, {'ph': placeholders})

    # 3. placeholder IATAs key on the SF Top Account
    frappe.db.sql("""
        UPDATE `tabMIDT Record` m
        JOIN `tabSabre Agency` a ON a.sf_top_account = m.sf_top_account
        SET m.agency = a.name
        WHERE m.iata IN %(ph)s AND IFNULL(a.sf_top_account, '') != ''
    """, {'ph': placeholders})

    frappe.db.commit()

    total = frappe.db.sql("SELECT COUNT(*) FROM `tabMIDT Record`")[0][0]
    unresolved = frappe.db.sql(
        "SELECT COUNT(*) FROM `tabMIDT Record` WHERE agency IS NULL")[0][0]
    unresolved_vol = frappe.db.sql(
        "SELECT IFNULL(SUM(total_bookings), 0) FROM `tabMIDT Record` "
        "WHERE agency IS NULL")[0][0]

    return {'total': total, 'resolved': total - unresolved,
            'unresolved': unresolved, 'unresolved_bookings': int(unresolved_vol or 0)}


def resolve_agency_for(agency_name):
    """Re-resolve just one agency's rows. Cheap - indexed, one card's worth."""
    placeholders = ('1111111', '9999999')

    # Release whatever this card currently holds, then re-claim.
    frappe.db.sql("UPDATE `tabMIDT Record` SET agency = NULL WHERE agency = %s",
                  (agency_name,))

    frappe.db.sql("""
        UPDATE `tabMIDT Record` m
        JOIN `tabSabre Agency` a ON a.iata = m.iata
        SET m.agency = a.name
        WHERE a.name = %(n)s AND IFNULL(a.iata, '') != ''
          AND m.iata NOT IN %(ph)s AND m.agency IS NULL
    """, {'n': agency_name, 'ph': placeholders})

    frappe.db.sql("""
        UPDATE `tabMIDT Record` m
        JOIN `tabSabre Agency` a ON a.iata = m.iata AND a.sc_code = m.sc_code
        SET m.agency = a.name
        WHERE a.name = %(n)s AND IFNULL(a.sc_code, '') NOT IN ('', '-', 'UNK')
          AND m.iata NOT IN %(ph)s
    """, {'n': agency_name, 'ph': placeholders})

    frappe.db.sql("""
        UPDATE `tabMIDT Record` m
        JOIN `tabSabre Agency` a ON a.sf_top_account = m.sf_top_account
        SET m.agency = a.name
        WHERE a.name = %(n)s AND m.iata IN %(ph)s
          AND IFNULL(a.sf_top_account, '') != '' AND m.agency IS NULL
    """, {'n': agency_name, 'ph': placeholders})


def resolve_agency_on_save(doc, method=None):
    """doc_events hook. Stands down during an import - the import resolves
    everything in one pass at the end, and _sync_agencies creates ~141 cards."""
    if frappe.flags.get('in_midt_import'):
        return
    try:
        resolve_agency_for(doc.name)
    except Exception:
        # Never block saving an agency because resolution failed; the nightly
        # pass will pick it up.
        frappe.log_error(frappe.get_traceback(), 'resolve_agency_on_save')
